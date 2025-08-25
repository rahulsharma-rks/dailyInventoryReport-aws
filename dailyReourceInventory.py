import os
import json
import tempfile
import boto3
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# AWS clients
config = boto3.client('config')
s3 = boto3.client('s3')
ses = boto3.client('ses')
cloudtrail = boto3.client('cloudtrail')

# Environment variables
S3_BUCKET = os.environ['REPORT_S3_BUCKET']
EMAIL_FROM = os.environ['EMAIL_FROM']
EMAIL_TO = os.environ['EMAIL_TO']

def get_user_from_cloudtrail(resource_id, event_time):
    """Get IAM user who performed the action from CloudTrail"""
    try:
        start_time = event_time - timedelta(minutes=5)
        end_time = event_time + timedelta(minutes=5)
        
        response = cloudtrail.lookup_events(
            LookupAttributes=[
                {
                    'AttributeKey': 'ResourceName',
                    'AttributeValue': resource_id
                }
            ],
            StartTime=start_time,
            EndTime=end_time,
            MaxItems=1
        )
        
        if response['Events']:
            return response['Events'][0].get('Username', 'Unknown')
    except Exception:
        pass
    return 'Unknown'

def get_resource_details(resource_type, resource_id, region):
    """Get additional resource details based on type"""
    details = {}
    try:
        if resource_type == 'AWS::EC2::Instance':
            ec2 = boto3.client('ec2', region_name=region)
            response = ec2.describe_instances(InstanceIds=[resource_id])
            if response['Reservations']:
                instance = response['Reservations'][0]['Instances'][0]
                details['InstanceType'] = instance.get('InstanceType', '')
                details['State'] = instance.get('State', {}).get('Name', '')
        elif resource_type == 'AWS::S3::Bucket':
            details['State'] = 'Active'
        elif resource_type == 'AWS::Lambda::Function':
            lambda_client = boto3.client('lambda', region_name=region)
            response = lambda_client.get_function(FunctionName=resource_id)
            details['Runtime'] = response['Configuration'].get('Runtime', '')
            details['State'] = response['Configuration'].get('State', '')
    except Exception:
        details['State'] = 'Unknown'
    return details

def lambda_handler(event, context):
    # Define yesterday's date range (UTC)
    today = datetime.utcnow().date()
    yesterday = today - timedelta(days=1)
    start_time = datetime(yesterday.year, yesterday.month, yesterday.day)
    end_time = datetime(today.year, today.month, today.day)

    # Query AWS Config for all resources and their history
    items = []
    
    # Get current resources
    paginator = config.get_paginator('select_resource_config')
    current_query = (
        "SELECT resourceId, resourceType, tags, awsRegion, "
        "configurationItemCaptureTime, configurationItemStatus, "
        "resourceCreationTime, configuration "
        "WHERE configurationItemStatus IN ('ResourceDiscovered', 'OK')"
    )
    
    for page in paginator.paginate(Expression=current_query):
        for r in page.get('Results', []):
            data = json.loads(r)
            
            # Parse configuration for additional details
            config_data = json.loads(data.get('configuration', '{}'))
            
            # Get resource details
            resource_details = get_resource_details(
                data.get('resourceType'),
                data.get('resourceId'),
                data.get('awsRegion')
            )
            
            # Parse tags
            tags = data.get('tags', {})
            tag_string = ', '.join([f"{k}:{v}" for k, v in tags.items()]) if tags else 'No tags'
            
            # Determine if this is a recent change
            capture_time = datetime.fromisoformat(data.get('configurationItemCaptureTime').replace('Z', '+00:00'))
            creation_time = data.get('resourceCreationTime')
            if creation_time:
                creation_time = datetime.fromisoformat(creation_time.replace('Z', '+00:00'))
            
            # Determine change type and color
            change_type = 'Existing'
            if creation_time and creation_time.date() == yesterday:
                change_type = 'Created'
            elif capture_time.date() == yesterday:
                change_type = 'Modified'
            
            # Get user information
            user = get_user_from_cloudtrail(data.get('resourceId'), capture_time)
            
            items.append({
                'IAM User': user,
                'Resource ID': data.get('resourceId'),
                'Resource Type': data.get('resourceType'),
                'Current State': resource_details.get('State', 'Active'),
                'Region': data.get('awsRegion'),
                'Creation Date': creation_time.strftime('%Y-%m-%d %H:%M:%S') if creation_time else 'Unknown',
                'Last Modified': capture_time.strftime('%Y-%m-%d %H:%M:%S'),
                'Change Type': change_type,
                'Tags': tag_string,
                'Additional Info': json.dumps(resource_details) if resource_details else ''
            })
    
    # Get deleted resources from yesterday
    deleted_query = (
        f"SELECT resourceId, resourceType, tags, awsRegion, "
        f"configurationItemCaptureTime, resourceDeletionTime "
        f"WHERE configurationItemStatus = 'ResourceDeleted' "
        f"AND resourceDeletionTime BETWEEN '{start_time}' AND '{end_time}'"
    )
    
    for page in paginator.paginate(Expression=deleted_query):
        for r in page.get('Results', []):
            data = json.loads(r)
            deletion_time = datetime.fromisoformat(data.get('resourceDeletionTime').replace('Z', '+00:00'))
            
            tags = data.get('tags', {})
            tag_string = ', '.join([f"{k}:{v}" for k, v in tags.items()]) if tags else 'No tags'
            
            user = get_user_from_cloudtrail(data.get('resourceId'), deletion_time)
            
            items.append({
                'IAM User': user,
                'Resource ID': data.get('resourceId'),
                'Resource Type': data.get('resourceType'),
                'Current State': 'Deleted',
                'Region': data.get('awsRegion'),
                'Creation Date': 'Unknown',
                'Last Modified': deletion_time.strftime('%Y-%m-%d %H:%M:%S'),
                'Change Type': 'Deleted',
                'Tags': tag_string,
                'Additional Info': ''
            })

    # Create DataFrame
    if not items:
        items = [{
            'IAM User': 'N/A',
            'Resource ID': 'No changes detected',
            'Resource Type': 'N/A',
            'Current State': 'N/A',
            'Region': 'N/A',
            'Creation Date': 'N/A',
            'Last Modified': 'N/A',
            'Change Type': 'N/A',
            'Tags': 'N/A',
            'Additional Info': 'N/A'
        }]
    
    df = pd.DataFrame(items)
    
    # Generate Excel with formatting
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        temp_path = tmp.name
    
    # Create workbook and worksheet
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'AWS Resources Report'
    
    # Add data to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Apply color coding and formatting
    green_fill = PatternFill('solid', fgColor='C6EFCE')  # Created
    yellow_fill = PatternFill('solid', fgColor='FFEB9C') # Modified
    red_fill = PatternFill('solid', fgColor='FFC7CE')    # Deleted
    blue_fill = PatternFill('solid', fgColor='B6D7FF')   # Existing
    
    # Header formatting
    header_fill = PatternFill('solid', fgColor='366092')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Format data rows based on change type
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        change_type = ws[f'H{row_num}'].value  # Change Type column
        
        if change_type == 'Created':
            fill = green_fill
        elif change_type == 'Modified':
            fill = yellow_fill
        elif change_type == 'Deleted':
            fill = red_fill
        else:
            fill = blue_fill
        
        for cell in row:
            cell.fill = fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(temp_path)
    
    # Upload to S3
    key = f"reports/{yesterday.isoformat()}-aws-resource-report.xlsx"
    s3.upload_file(temp_path, S3_BUCKET, key)
    
    # Generate summary statistics
    summary = df['Change Type'].value_counts().to_dict()
    summary_text = '\n'.join([f"{k}: {v}" for k, v in summary.items()])
    
    # Send email via SES
    url = s3.generate_presigned_url(
        'get_object',
        Params={'Bucket': S3_BUCKET, 'Key': key},
        ExpiresIn=86400
    )
    
    email_body = f"""AWS Resource Daily Report - {yesterday.isoformat()}

Summary of Changes:
{summary_text}

Total Resources Tracked: {len(df)}

Color Coding:
🟢 Green: Newly Created Resources
🟡 Yellow: Modified Resources  
🔴 Red: Deleted Resources
🔵 Blue: Existing Resources (no changes)

Download your detailed Excel report here:
{url}

This link expires in 24 hours.

Report generated at: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"""
    
    ses.send_email(
        Source=EMAIL_FROM,
        Destination={'ToAddresses': [EMAIL_TO]},
        Message={
            'Subject': {'Data': f"📊 AWS Resource Report - {yesterday.isoformat()}"},
            'Body': {'Text': {'Data': email_body}}
        }
    )
    
    # Clean up temp file
    os.unlink(temp_path)
    
    return {
        'statusCode': 200,
        'body': json.dumps({
            'status': 'SUCCESS',
            'report_key': key,
            'resources_tracked': len(df),
            'summary': summary
        })
    }
